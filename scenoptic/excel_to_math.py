from __future__ import annotations

import more_itertools
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass
from enum import Enum
from itertools import takewhile, count
from more_itertools import split_before
from openpyxl import load_workbook
from pathlib import Path
from typing import Optional, MutableMapping, Sequence, Mapping, Callable, Tuple, TypedDict, Literal, Union, \
    Iterable, Iterator

from codegen.utils import visitor_for
from math_rep.expr import FormalContent, Term
from scenoptic.scenoptic_expr import Cell, Cells
from math_rep.expression_types import M_NUMBER, QualifiedName
from scenoptic.excel_data import CellType, BooleanCell, IntegerCell, NonNegIntParser, FloatCell, NonNegFloatParser, \
    StringCell, INTEGER_CELL_TYPE, CellTypeFactory, ExcelData
from scenoptic.parse_excel import parse_excel_cell_or_range
from scenoptic.xl_utils import xl_range_elements, cell_to_coords, coords_to_cell, \
    get_sheet_in_cell_qn_as_string, get_cell_qn_as_excel_string, xl_cell_or_range_elements_qn


class OptimizationDirection(Enum):
    MAXIMIZE = True
    MINIMIZE = False


MAXIMIZE_WORDS = ('max', 'maximize')
MINIMIZE_WORDS = ('min', 'minimize')


class AbstractConstraint:
    pass


class Constant(AbstractConstraint):
    def __init__(self, value):
        self.value = value


@dataclass
class Constraint(AbstractConstraint):
    code: FormalContent
    predecessors: Sequence[QualifiedName]


@visitor_for(FormalContent)
class AbstractExprVisitor:
    pass


class CollectCells(AbstractExprVisitor):
    def visit_cell(self, cell: Cell):
        return cell

    def visit_cells(self, cells: Cells):
        return list(cells)

    def collect(self, term: Term) -> Iterator[Cell]:
        return more_itertools.collapse(self.visit(term))


CELL_COLLECTOR = CollectCells()

CELL_TYPES_PARSERS = {'int': IntegerCell, 'float': FloatCell, 'bool': BooleanCell, 'string': StringCell,
                      'int+': NonNegIntParser, 'float+': NonNegFloatParser}

CELL_TYPES = dict(int=IntegerCell, float=FloatCell, bool=BooleanCell, string=StringCell)

DEFAULT_CELL_TYPE_PARSER_FACTORY = CellTypeFactory(CELL_TYPES)


class ExcelError:
    pass


EXCEL_ARG_TYPE_TYPE = Literal['scalar', 'array', 'both']
EXCEL_BLANK_TREATMENT_TYPE = Literal['ignore', 'include', 0, '']


@dataclass
class ExcelOperator:
    # is the result of the operator a scalar or an array?
    scalar_result: bool
    # is argument a scalar or an array?  (no distinction yet between ranges and arrays, etc.)
    argument_type: Callable[[int], EXCEL_ARG_TYPE_TYPE]
    # treatment of blank cells
    blank_treatment: EXCEL_BLANK_TREATMENT_TYPE
    # must the sizes of all array argument match exactly? (in what we support)
    ranges_must_match: bool = True
    # can arrays be given to scalar args to make scalar-result operator into an array formula?
    can_be_used_as_array_formula: bool = True
    # if an array operator is 1:1, function that transforms input indexes to output indexes; otherwise None
    one_to_one: Optional[Callable[[Tuple[int, ...]], Tuple[int, ...]]] = None
    # for array operators, the size of the output from the size of the input if available, otherwise None
    output_size: Optional[Callable[[Tuple[int, ...]], Tuple[int, ...]]] = None
    # if operator accepts "criteria" arguments, an iterable that contains their indexes
    criteria_arguments: Iterable[int] = ()


def scalar_operator(_: int) -> Literal['scalar']:
    return 'scalar'


def permissive_operator(_: int) -> Literal['both']:
    return 'both'


def range_operator(_: int) -> Literal['array']:
    return 'array'


ARITHMETIC_OPERATOR = ExcelOperator(True, scalar_operator, 0)

EXCEL_SUM_OPERATORS = 'SUM,SUMIF,SUMIFS'.split(',')
EXCEL_COUNT_OPERATORS = 'COUNT,COUNTIF,COUNTIFS'.split(',')
EXCEL_AGGREGATE_OPERATORS = EXCEL_SUM_OPERATORS + EXCEL_COUNT_OPERATORS


def make_iterable(func: Callable[[], Iterator]) -> Iterable:
    class CustomIterable(Iterable):
        def __iter__(self):
            return func()

    return CustomIterable()


# N.B. apply rule ExpandExcelSumMultipleArgs before using SUM operator!
EXCEL_OPERATORS = {'+': ARITHMETIC_OPERATOR,
                   '-': ARITHMETIC_OPERATOR,
                   '*': ARITHMETIC_OPERATOR,
                   '/': ARITHMETIC_OPERATOR,
                   '^': ARITHMETIC_OPERATOR,
                   '&': ExcelOperator(True, scalar_operator, ''),
                   'SUM': ExcelOperator(True, range_operator, 'ignore'),
                   'SUMIF': ExcelOperator(True, range_operator, 'ignore', criteria_arguments=[1]),
                   'SUMIFS': ExcelOperator(True, range_operator, 'ignore',
                                           criteria_arguments=make_iterable(lambda: count(2, 2))),
                   'COUNT': ExcelOperator(True, range_operator, 'ignore'),
                   'COUNTIF': ExcelOperator(True, range_operator, 'ignore', criteria_arguments=[1]),
                   'COUNTIFS': ExcelOperator(True, range_operator, 'ignore',
                                             criteria_arguments=make_iterable(lambda: count(1, 2)))}

EXTERNAL_LINK_PATTERN = r"""(^'\[\d+\].*)"""
EXTERNAL_LINK_RE = re.compile(EXTERNAL_LINK_PATTERN)


class Scenario(ExcelData, ABC):
    # TODO: infer default type for constants from their values?
    # FIXME!!! remove default_type, not used, require types for all cells
    def __init__(self, excel_file, sheet: str = None, default_type=M_NUMBER,
                 cell_type_parser_factory=DEFAULT_CELL_TYPE_PARSER_FACTORY):
        self.wb = load_workbook(Path(excel_file).resolve())
        if sheet:
            try:
                self.ws = self.wb[sheet]
            except KeyError:
                self.ws = self.wb.active
        else:
            self.ws = self.wb.active
        self.cell_type_parser_factory = cell_type_parser_factory
        self.named_ranges = {}
        self.parameters = []
        self.objectives = {}
        self.default_type = default_type
        self.types: MutableMapping[str, CellType] = {}  # defaultdict(lambda: self.default_type)
        self.warnings = []
        self.dummy_cells = {}
        self._populate_named_ranges()

    @abstractmethod
    def convert_to_language(self, expression: FormalContent):
        """
        Takes an abstract expression
        and converts it to target langauge
        """

    def _populate_named_ranges(self):
        # print(f'Defined Names :\n {self.wb.defined_names.definedName}')
        # print(f'External Links:\n {self.wb._external_links}')

        defined_names = {dn.name: (dn.localSheetId, dn.attr_text) for dn in self.wb.defined_names}
        for dn_name, args in defined_names.items():
            # FIXME: For private names, handle scenario where the same private named
            #  range VALUE appears in more than one sheet

            # FIXME!!! support named ranges that refers to external files,
            #  e.g.
            #   the following is named range `Stockout`
            #     name='Stockouts', comment=None, customMenu=None, description=None, help=None,
            #     statusBar=None, localSheetId=None, hidden=None, function=None, vbProcedure=None, xlm=None,
            #     functionGroupId=None, shortcutKey=None, publishToServer=None,
            #     workbookParameter=None,
            #     attr_text="'[1]Business Model Specification'!$F$3:$F$15"]
            #  Notice, `attr_text` starts with reference to a file `[1]`
            #   See: https://stackoverflow.com/questions/34096323/identify-external-workbook-links-using-openpyxl
            if args[0] is not None:
                # Private named ranges scoped to the `localSheetId` attribute of named range
                dest = self.wb.defined_names.get(dn_name, scope=args[0]).destinations
            else:
                try:
                    dest = self.wb.defined_names[dn_name].destinations
                except KeyError:
                    print(f'Loading Private named range going through exception {dn_name}')
                    dest = self.wb.defined_names.get(dn_name, scope=args[0]).destinations
            try:
                for sheet_name, coords in dest:
                    m = EXTERNAL_LINK_RE.match(args[1])
                    if m is not None:
                        self.warn(f'We do not support named ranges with external links "{args[1]}"')
                        continue

                    value = parse_excel_cell_or_range(args[1], self)
                    self.named_ranges[dn_name] = dict(cells=coords, value=value)
                    # print(f'dn={dn_name} sheet={sheet_name} cells={coords} -> {args[1]} --> {value}')
            except AttributeError as e:
                # Ignore elements that may have errors in them:
                # e.g.
                #   name='solver_ver', comment=None, customMenu=None, description=None, help=None,
                #   statusBar=None, localSheetId=1, hidden=True, function=None, vbProcedure=None,
                #   xlm=None, functionGroupId=None, shortcutKey=None, publishToServer=None,
                #   workbookParameter=None, attr_text='integer']
                #
                #  In the above, `solver_var` is a defined name range, which does not appear in the excel `Name Manager`
                #  and it refers to sheet `localSheetId=1` but the sheet value is `attr_text='3' `  which is
                #  bad formatted and results with exception AttributeError...
                self.warn(f'Named Range "{dn_name}" with illegal range expression in "attr_text"="{args[1]}"')
                pass

    def is_named_range(self, named_range_candidate: str) -> bool:
        return self.named_ranges.get(named_range_candidate, None) is not None

    def named_range_sheet(self, named_range_candidate: str) -> Union[str, None]:
        if dn := self.named_ranges.get(named_range_candidate):
            return dn['value'].sheet
        return None

    def named_range_cells(self, named_range_candidate: str) -> Union[str, None]:
        if dn := self.named_ranges.get(named_range_candidate):
            return dn['cells']
        return None

    def named_range_value(self, named_range_candidate: str) -> Union[str, None]:
        if dn := self.named_ranges.get(named_range_candidate):
            return dn['value']
        return None

    def add_parameters(self, parameters: str, ctype: CellType):
        self.parameters.extend(xl_cell_or_range_elements_qn(parameters, self.specification_sheet_name()))
        self.set_type(parameters, ctype)

    def add_constraint(self, target_cells: str):
        """
        Add a boolean constraint.

        :param target_cells: cell or range containing constraint(s)
        """
        self.add_objective(target_cells, OptimizationDirection.MINIMIZE, level=1, ctype=INTEGER_CELL_TYPE)

    def add_objective(self, target_cells: str, direction: OptimizationDirection, level: int = 0,
                      ctype: Optional[CellType] = None):
        """
        Add an objective.

        Positive levels are hard constraints; zero or negative levels are soft constraints (i.e., objectives).
        Higher levels are stronger constraints.  The default

        :param target_cells: cell or range containing value of objective(s)
        :param direction: OptimizationDirection.MAXIMIZE or OptimizationDirection.MINIMIZE, as appropriate
        :param level: level of constraint
        :param ctype: optional cell type
        """
        # Note: objectives are keyed on strings, not QNs!
        self.objectives[target_cells] = (level, direction)
        if ctype:
            self.set_type(target_cells, ctype)

    def set_type(self, cells: Union[str, QualifiedName], ctype: CellType):
        """
        Define the type of a cell or range

        :param cells: cell or range
        :param ctype: type of cells
        """
        if isinstance(cells, QualifiedName):
            self.types[cells] = ctype
        else:
            for cell in xl_cell_or_range_elements_qn(cells, self.specification_sheet_name()):
                self.types[cell] = ctype

    def warn(self, msg):
        self.warnings.append(msg)
        # FIXME: remove printing here
        print(f'WARNING: {msg}')

    def cell_value(self, row, col, req_type=None, sheet=None, cast_to=None, empty_values=()):
        value = (self.wb[sheet] if sheet is not None else self.ws).cell(row, col).value
        if value is None or value == '' or isinstance(value, str) and value in empty_values:
            return None
        if req_type and not isinstance(value, req_type):
            return None
        if cast_to is not None:
            return cast_to(value)
        return value

    def cell_qn_value(self, cell: QualifiedName, req_type=None, cast_to=None, empty_values=()):
        """
        Get the value of the QualifiedName cell .  If ``req_type`` is supplied and the type of the value
        is not on ``req_type``, or if the value is on ``empty_values``, return None.

        :param cell:QualifiedName cell
        :param req_type: sequence of types or None
        :param cast_to: type to cast the result to
        :param empty_values: sequence of values indicating empty cell
        :return: cell value or None
        """
        if (value := self.dummy_cells.get(cell)) is None:
            sheet = get_sheet_in_cell_qn_as_string(cell) or self.specification_sheet_name()
            normalized_cell_name = get_cell_qn_as_excel_string(cell)
            if normalized_cell_name is None:
                return None
            value = self.wb[sheet][normalized_cell_name].value
        if value is None or value == '' or isinstance(value, str) and value in empty_values:
            return None
        if req_type and not isinstance(value, req_type):
            return None
        if cast_to is not None:
            return cast_to(value)
        return value

    def _get_cells_or_range_if_named_range(self, named_range_candidate: str) -> str:
        if self.is_named_range(named_range_candidate):
            return f"'{self.named_range_sheet(named_range_candidate)}'!" \
                   f'{self.named_range_cells(named_range_candidate)}'
        else:
            return named_range_candidate

    def _parse_type(self, row, col) -> Optional[CellType]:
        """
        Parse a type whose arguments start in ``(row, col)``

        :return: CellType or None if can't parse
        """
        type_name = self.cell_value(row, col, str).strip()
        optional = False
        if type_name[0] == '(' and type_name[-1] == ')':
            type_name = type_name[1:-1].strip()
            optional = True
        parser = CELL_TYPES_PARSERS.get(type_name)
        if parser is None:
            self.warn(f'Unknown cell type: {type_name} in {coords_to_cell(row, col)}')
            return None
        return parser.parse_type_args(row, col, self, self.cell_type_parser_factory, optional=optional)

    # FIXME! add check for valid cell/range in parsers
    def _parse_decisions(self, row, first_col):
        """
        Parse a decision-variable specification, which consists of a cell or range followed by a type

        :param row: number of row containing decision-variable specification
        :param first_col: number of column containing specifications (one after header)
        """
        param_cells = self.cell_value(row, first_col, str)
        if param_cells is None:
            self.warn(f'Bad parameter cell or range {param_cells} in {coords_to_cell(row, first_col)}')
            return
        param_cells = self._get_cells_or_range_if_named_range(param_cells)
        ctype = self._parse_type(row, first_col + 1)
        if ctype is None:
            self.warn(f'Bad type for {param_cells} in {coords_to_cell(row, first_col + 1)}')
            return
        self.add_parameters(param_cells, ctype)

    def _parse_objectives(self, row, first_col):
        """
        Parse an objective specification, which consists of a cell or range, followed by an optional level number,
        followed by an optimization direction (minimize or maximize), followed by a type

        :param row: number of row containing objective specification
        :param first_col: number of column containing specifications (one after header)
        """
        objective_cells = self.cell_value(row, first_col, str)
        if objective_cells is None:
            self.warn(f'Bad objective cell or range {objective_cells}')
            return
        objective_cells = self._get_cells_or_range_if_named_range(objective_cells)
        objective_level_or_direction = self.cell_value(row, first_col + 1, (int, str))
        if objective_level_or_direction is None:
            self.warn(f'Bad objective level or direction {objective_level_or_direction}')
            return
        if isinstance(objective_level_or_direction, int):
            objective_level = objective_level_or_direction
            type_col = first_col + 3
            objective_direction_str = self.cell_value(row, first_col + 2, str)
        else:
            objective_level = 0
            type_col = first_col + 2
            objective_direction_str = objective_level_or_direction
        if not isinstance(objective_direction_str, str):
            self.warn(f'Objective direction must be a string: {objective_direction_str}')
            return
        if objective_direction_str.lower() in MAXIMIZE_WORDS:
            objective_direction = OptimizationDirection.MAXIMIZE
        elif objective_direction_str.lower() in MINIMIZE_WORDS:
            objective_direction = OptimizationDirection.MINIMIZE
        else:
            self.warn(f'Unknown objective direction {objective_direction_str}')
            return
        ctype = self._parse_type(row, type_col)
        if ctype is None:
            self.warn(f'Bad type for {objective_cells} in {coords_to_cell(row, type_col)}')
            return
        self.add_objective(objective_cells, objective_direction, objective_level)
        self.set_type(objective_cells, ctype)

    def _parse_constraints(self, row, first_col):
        """
        Parse a constraint specification, which consists of a cell or range

        :param row: number of row containing constraint specification
        :param first_col: number of column containing specifications (one after header)
        """
        constraint_cells = self.cell_value(row, first_col, str)
        if constraint_cells is None:
            self.warn(f'Bad constraint cell or range {constraint_cells}')
            return
        constraint_cells = self._get_cells_or_range_if_named_range(constraint_cells)
        self.add_constraint(constraint_cells)

    def _parse_types(self, row, first_col):
        """
        Parse a type specification, which consists of a cell or range followed by a type

        :param row: number of row containing type specification
        :param first_col: number of column containing specifications (one after header)
        """
        type_cells = self.cell_value(row, first_col, str)
        if type_cells is None:
            self.warn(f'Bad cell or range {type_cells}')
            return
        type_cells = self._get_cells_or_range_if_named_range(type_cells)
        ctype = self._parse_type(row, first_col + 1)
        if ctype is None:
            self.warn(f'Bad type for {type_cells} in {coords_to_cell(row, first_col + 1)}')
            return
        self.set_type(type_cells, ctype)

    def _block_parsers(self) -> Mapping[str, Callable[[Sequence[int], int], None]]:
        """
        Return a mapping from headers to parsing functions for the corresponding definition block in the spreadsheet.

        The functions should be bound methods of ``self`` so that they can add information to ``self``.

        Override this method to provide more block types.
        """
        param_parser = self._parse_decisions
        objective_parser = self._parse_objectives
        constraint_parser = self._parse_constraints
        type_parser = self._parse_types
        return dict(parameters=param_parser, parameter=param_parser, decision=param_parser, decisions=param_parser,
                    objective=objective_parser, objectives=objective_parser,
                    constraint=constraint_parser, constraints=constraint_parser,
                    type=type_parser, types=type_parser)

    def sheet_names(self):
        return self.wb.sheetnames

    def specification_sheet_name(self):
        return self.ws.title

    def find_translation_arguments(self):
        """
        Find in the spreadsheet a section containing the arguments for the translation.

        The section must start with a cell containing "Scenoptic" in a cell inside the range ``A1:Z200``.
        The column below that cell contains the argument names (see below), and the following columns hold the
        values.  In some cases, the argument is a list that continues in the following rows (no argument name
        necessary).  All strings are not case sensitive.

        The list of arguments ends at the first row that has the first two columns blank.  Rows that have a value
        that starts with a hyphen in the first column are ignored.

        The arguments are:
            - ``name``: the base name for classes in this scenario
            - ``package``: the Java base package name
            - ``parameter[s]``: a list of cells or ranges containing the decisions the solver needs to make: the cell or cell-range, and the type (including parameters)
            - ``objective[s]``: a list of the objectives: the cell or cell-range; the objective level (optional); one of ``maximize``, ``max``, ``minimize``, or ``min``; and the type (including parameters)
            - ``constraint[s]``: a list of constraints, which are boolean objectives to be maximized, at level 1.
            - ``types``: a list of cells or ranges with the corresponding type (including parameters).
        """

        def is_comment(value):
            return isinstance(value, str) and ((canon_value := value.lower()).startswith('-')
                                               or canon_value.startswith("'-"))

        arg_header = None
        for cell in xl_range_elements(1, 1, self.ws.max_column, self.ws.max_row):
            header = self.ws[cell].value
            if isinstance(header, str) and header.lower() == 'scenoptic':
                arg_header = cell
                break
        if arg_header is None:
            return
        arg_row, header_col = cell_to_coords(arg_header)
        first_arg_col = header_col + 1
        rows = filter(lambda row: not is_comment(self.cell_value(row, header_col)),
                      takewhile(lambda row: self.cell_value(row, header_col) is not None
                                            or self.cell_value(row, first_arg_col) is not None,
                                range(arg_row + 1, self.ws.max_row + 1)))
        groups = split_before(rows, lambda row: self.cell_value(row, header_col) is not None)
        for group in groups:
            first_row = group[0]
            orig_header = self.cell_value(first_row, header_col)
            header = orig_header.lower()
            parser = self._block_parsers().get(header)
            if parser is None:
                self.warn(f'Unknown argument header: {orig_header}')
            else:
                for row in group:
                    parser(row, first_arg_col)

    def validate(self) -> Sequence[ExcelError]:
        """
        Check that the problem is well defined.

        Checks that:
            - There are no parsing errors in the translation argument table
            - All objectives are int or bool (after propagation)
            - All provided types match inferred types
            - There is at least one constraint or objective
            - ...


        Override this method to add more checks.

        :return: sequence of errors, empty if all checks passed
        """
        # FIXME!! fill in
        return ()

    @abstractmethod
    def validate_initialization(self):
        """
        Raise exception if any of the input data and parameters is not properly defined
        """

    def get_type(self, cell: QualifiedName) -> Optional[CellType]:
        return self.types.get(cell)


class CellSpec(TypedDict):
    sheet: str
    row: int
    col: int


def extract_cell_components(name: str) -> Tuple[str, int, int]:
    sep = str.rindex(name, '!')
    row, col = cell_to_coords(name[sep + 1:])
    return name[:sep], row, col


def extract_cell_components_to_dict(name: str) -> CellSpec:
    sheet, row, col = extract_cell_components(name)
    return dict(sheet=sheet, row=row, col=col)


if __name__ == '__main__':
    print(extract_cell_components_to_dict('v2!j16'))
